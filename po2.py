import tarfile
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# Base directory (script can run from anywhere)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
AFTER_DIR = BASE_DIR / "After_Run"
BEFORE_DIR = BASE_DIR / "Before_Run"


# =========================
# Dynamic PSV filename patterns (Context 2)
# =========================
cnb_list = [
    "error_summary_cnb_out_*.psv",
    "summary_count_cnb_out_*.psv",
]

ccms_list = [
    "error_summary_ccms_out_*.psv",
    "summary_ccms_out_*.psv",
    "summary_count_ccms_out_*.psv",
]

cms_list = [
    "error_summary_cms_out_*.psv",
    "summary_cms_out_*.psv",
    "summary_count_cms_out_*.psv",
]


# =========================
# TAR prefixes (timestamp part is dynamic)
# =========================
CNB_OUTER_PREFIX = "cnb_in_out"
CNB_INNER_OUT_PREFIX = "cnb_out"

CCMS_OUTER_PREFIX = "lgd_ccms_in_out"
CCMS_INNER_OUT_PREFIX = "ccms_out"

COMM_OUTER_PREFIX = "lgd_commercial_in_out"
CMS_INNER_OUT_PREFIX = "cms_out"
ESN_INNER_OUT_PREFIX = "esn_out"


# =========================
# Excel Layout Rules
# =========================
COL_GAP_BETWEEN_BLOCKS = 2  # 2 blank columns between After and Before blocks
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow highlight for diffs
TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# If you want EXACT section start feel like screenshot for 3-section reports:
# Error Summary label around row 2, Summary label around row 10, Summary count label around row 20
MIN_SECTION_LABEL_ROWS_FOR_3_TABLES = [2, 10, 20]


# =========================
# Robust File Find Helpers
# =========================
TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz")


def find_by_prefix(folder: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    candidates = [p for p in folder.iterdir()
                  if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes)]

    if not candidates:
        present = sorted([p.name for p in folder.iterdir() if p.is_file()])
        raise FileNotFoundError(
            f"Couldn't find any file starting with '{prefix}' under: {folder}\n"
            f"Files present: {present}"
        )

    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"⚠️ Multiple matches for '{prefix}' in {folder}. Using newest: {candidates[0].name}")
    return candidates[0]


def find_inside_extracted(root: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    matches = [p for p in root.rglob("*")
               if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes)]

    if not matches:
        raise FileNotFoundError(f"Couldn't find inner tar starting with '{prefix}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple inner matches for '{prefix}'. Using newest: {matches[0].name}")
    return matches[0]


def find_psv_by_glob(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Couldn't find PSV '{pattern}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple PSV matches for '{pattern}'. Using newest: {matches[0].name}")
    return matches[0]


def extract_tar(tar_path: Path, extract_to: Path) -> None:
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(path=extract_to)


# =========================
# PSV Parsing (CRITICAL FIX)
# =========================
def read_psv_preserve_shape(psv_path: Path) -> pd.DataFrame:
    """
    Reads a PSV file by splitting each line on literal '|', pads rows to max field count.
    This preserves:
      - metadata rows like "Business Date: ..." (single cell row)
      - real table header rows like "Records|SingleName|..."
      - key/value rows like "Total Records|38118"
    We do NOT promote any row to dataframe columns.
    """
    lines = psv_path.read_text(encoding="utf-8", errors="replace").splitlines()
    rows = [line.split("|") for line in lines]
    max_len = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    # Use numeric columns; writing to excel won't add extra headers
    return pd.DataFrame(rows)


def load_psv_dfs_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, psv_patterns):
    """
    - find outer tar by prefix in run_folder
    - extract it
    - find inner tar(s) by prefix inside extracted outer
    - extract all inner tar(s)
    - search PSV patterns inside extracted inner content
    - return list of dataframes
    """
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        dfs = []
        for pat in psv_patterns:
            psv_file = find_psv_by_glob(inner_extract_root, pat)
            dfs.append(read_psv_preserve_shape(psv_file))

        return dfs
    finally:
        tmpdir.cleanup()


# =========================
# Excel writing helpers (NO table coloring)
# =========================
def autosize_columns(ws, col_start: int, col_end: int, row_start: int, row_end: int, min_w=10, max_w=60):
    for c in range(col_start, col_end + 1):
        max_len = 0
        for r in range(row_start, row_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, max_len + 2), max_w)


def write_title(ws, row: int, col: int, text: str, span_cols: int):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_LEFT
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span_cols - 1)


def write_section_label(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = ALIGN_LEFT


def write_df_values_only(ws, df: pd.DataFrame, start_row: int, start_col: int):
    """
    Write dataframe values ONLY (no header row).
    Returns (top, left, bottom, right) range.
    """
    nrows, ncols = df.shape
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + i, column=start_col + j, value=df.iat[i, j])

    top = start_row
    left = start_col
    bottom = start_row + nrows - 1
    right = start_col + ncols - 1
    return top, left, bottom, right


def compare_and_highlight(ws, rng_after, rng_before):
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    a_rows = a_bottom - a_top + 1
    a_cols = a_right - a_left + 1
    b_rows = b_bottom - b_top + 1
    b_cols = b_right - b_left + 1

    max_rows = max(a_rows, b_rows)
    max_cols = max(a_cols, b_cols)

    for r in range(max_rows):
        for c in range(max_cols):
            a_exists = (r < a_rows and c < a_cols)
            b_exists = (r < b_rows and c < b_cols)

            a_r, a_c = a_top + r, a_left + c
            b_r, b_c = b_top + r, b_left + c

            a_val = ws.cell(a_r, a_c).value if a_exists else ""
            b_val = ws.cell(b_r, b_c).value if b_exists else ""

            a_str = "" if a_val is None else str(a_val)
            b_str = "" if b_val is None else str(b_val)

            if a_str != b_str:
                if a_exists:
                    ws.cell(a_r, a_c).fill = DIFF_FILL
                if b_exists:
                    ws.cell(b_r, b_c).fill = DIFF_FILL


def build_validation_excel(
    out_file: str,
    sheet_name: str,
    after_tables: list,
    before_tables: list,
    section_names: list,
    after_block_title: str,
    before_block_title: str,
):
    """
    Places tables side-by-side (After on left, Before on right).
    - No table fill colors (only diff highlights).
    - For 3-table reports: section label rows align like screenshot (2,10,20) at minimum.
    - For 2-table reports: section label rows are dynamic.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Determine block width = max cols across all tables (After + Before)
    all_tables = after_tables + before_tables
    block_width = max((df.shape[1] for df in all_tables), default=1)

    left_start_col = 1  # A
    right_start_col = left_start_col + block_width + COL_GAP_BETWEEN_BLOCKS  # => if After ends at G, Before starts at J

    # Titles row 1
    write_title(ws, 1, left_start_col, after_block_title, block_width)
    write_title(ws, 1, right_start_col, before_block_title, block_width)

    # Decide label rows
    if len(section_names) == 3:
        min_label_rows = MIN_SECTION_LABEL_ROWS_FOR_3_TABLES
    else:
        min_label_rows = None

    after_ranges = []
    before_ranges = []

    # Write After block
    current_label_row = 2
    for idx, (sec_name, df) in enumerate(zip(section_names, after_tables)):
        if min_label_rows:
            current_label_row = max(current_label_row, min_label_rows[idx])

        write_section_label(ws, current_label_row, left_start_col, sec_name)
        # Table starts one row below section label
        rng = write_df_values_only(ws, df, current_label_row + 1, left_start_col)
        after_ranges.append(rng)

        # Next label row (default spacing): leave 3 blank rows between tables (can be adjusted)
        current_label_row = rng[2] + 1 + 3

    # Write Before block
    current_label_row = 2
    for idx, (sec_name, df) in enumerate(zip(section_names, before_tables)):
        if min_label_rows:
            current_label_row = max(current_label_row, min_label_rows[idx])

        write_section_label(ws, current_label_row, right_start_col, sec_name)
        rng = write_df_values_only(ws, df, current_label_row + 1, right_start_col)
        before_ranges.append(rng)

        current_label_row = rng[2] + 1 + 3

    # Highlight diffs (Context 7)
    for i in range(min(len(after_ranges), len(before_ranges))):
        compare_and_highlight(ws, after_ranges[i], before_ranges[i])

    # Autosize columns
    last_row = max(
        after_ranges[-1][2] if after_ranges else 1,
        before_ranges[-1][2] if before_ranges else 1,
    )
    autosize_columns(ws, left_start_col, left_start_col + block_width - 1, 1, last_row)
    autosize_columns(ws, right_start_col, right_start_col + block_width - 1, 1, last_row)

    wb.save(out_file)


# =========================
# Main Execution
# =========================
def main():
    # ---------------- CNB (2 tables) ----------------
    df_ar_cnb_es, df_ar_cnb_sc = load_psv_dfs_from_run(
        AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )
    df_br_cnb_es, df_br_cnb_sc = load_psv_dfs_from_run(
        BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )

    build_validation_excel(
        out_file="CNB_Validation.xlsx",
        sheet_name="CNB_Validation",
        after_tables=[df_ar_cnb_es, df_ar_cnb_sc],
        before_tables=[df_br_cnb_es, df_br_cnb_sc],
        section_names=["Error Summary", "Summary count"],
        after_block_title="After(Test build)",
        before_block_title="Before(prod)",
    )

    # ---------------- CCMS (3 tables) ----------------
    df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc = load_psv_dfs_from_run(
        AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )
    df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )

    build_validation_excel(
        out_file="CCMS_Validation.xlsx",
        sheet_name="CCMS_Validation",
        after_tables=[df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc],
        before_tables=[df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc],
        section_names=["Error Summary", "Summary", "Summary count"],
        after_block_title="After(Test build)",
        before_block_title="Before(prod)",
    )

    # ---------------- CMS (3 tables) ----------------
    # Need to extract both cms_out* and esn_out* from commercial tar
    df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc = load_psv_dfs_from_run(
        AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )
    df_br_cms_es, df_br_cms_sc, df_br_cms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )

    build_validation_excel(
        out_file="CMS_Validation.xlsx",
        sheet_name="CMS_Validation",
        after_tables=[df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc],
        before_tables=[df_br_cms_es, df_br_cms_sc, df_br_cms_scc],
        section_names=["Error Summary", "Summary", "Summary count"],
        after_block_title="After(Test build)",
        before_block_title="Before(prod)",
    )

    print("✅ Created files:")
    print(" - CNB_Validation.xlsx")
    print(" - CCMS_Validation.xlsx")
    print(" - CMS_Validation.xlsx")
    print("✅ Differences highlighted in light yellow (no other coloring).")


if __name__ == "__main__":
    main()