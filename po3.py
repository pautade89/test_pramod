import tarfile
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# Base directory (script can run from anywhere)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
AFTER_DIR = BASE_DIR / "After_Run"
BEFORE_DIR = BASE_DIR / "Before_Run"


# =========================
# Dynamic PSV filename patterns (Context 2)
# =========================
cnb_list = [
    "error_summary_cnb_out_*.psv",
    "summary_count_cnb_out_*.psv",
]

ccms_list = [
    "error_summary_ccms_out_*.psv",
    "summary_ccms_out_*.psv",
    "summary_count_ccms_out_*.psv",
]

cms_list = [
    "error_summary_cms_out_*.psv",
    "summary_cms_out_*.psv",
    "summary_count_cms_out_*.psv",
]


# =========================
# TAR prefixes (timestamp part is dynamic)
# =========================
CNB_OUTER_PREFIX = "cnb_in_out"
CNB_INNER_OUT_PREFIX = "cnb_out"

CCMS_OUTER_PREFIX = "lgd_ccms_in_out"
CCMS_INNER_OUT_PREFIX = "ccms_out"

COMM_OUTER_PREFIX = "lgd_commercial_in_out"
CMS_INNER_OUT_PREFIX = "cms_out"
ESN_INNER_OUT_PREFIX = "esn_out"


# =========================
# Excel Layout Rules (Reference format)
# =========================
AFTER_START_COL = 1   # Column A
BEFORE_MIN_START_COL = 10  # Column J (as in reference)
COL_GAP_BETWEEN_BLOCKS = 2  # 2 blank columns gap between blocks
ROW_GAP_BETWEEN_TABLES = 3  # 3 blank rows gap between sections

TITLE_ROW = 1  # "After_Run Results" and "Before_Run Results"

# For reference-like spacing:
# CCMS/CMS (3 tables): section label rows 3, 12, 19
# CNB (2 tables): section label rows 3, 12
REF_LABEL_ROWS_3 = [3, 12, 19]
REF_LABEL_ROWS_2 = [3, 12]

# Highlight fill for differences (Context 7)
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


# =========================
# Robust File Find Helpers
# =========================
TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz")


def find_by_prefix(folder: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    candidates = [
        p for p in folder.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes)
    ]

    if not candidates:
        present = sorted([p.name for p in folder.iterdir() if p.is_file()])
        raise FileNotFoundError(
            f"Couldn't find any file starting with '{prefix}' under: {folder}\n"
            f"Files present: {present}"
        )

    # newest file if multiple
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"⚠️ Multiple matches for '{prefix}' in {folder}. Using newest: {candidates[0].name}")
    return candidates[0]


def find_inside_extracted(root: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    matches = [
        p for p in root.rglob("*")
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes)
    ]
    if not matches:
        raise FileNotFoundError(f"Couldn't find inner tar starting with '{prefix}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple inner matches for '{prefix}'. Using newest: {matches[0].name}")
    return matches[0]


def find_psv_by_glob(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Couldn't find PSV '{pattern}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple PSV matches for '{pattern}'. Using newest: {matches[0].name}")
    return matches[0]


def extract_tar(tar_path: Path, extract_to: Path) -> None:
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(path=extract_to)


# =========================
# PSV Parsing (Preserve shape exactly)
# =========================
def read_psv_preserve_shape(psv_path: Path) -> pd.DataFrame:
    """
    Reads PSV as raw lines split by literal '|', pads rows to max column count.
    Does NOT treat any row as header.
    """
    lines = psv_path.read_text(encoding="utf-8", errors="replace").splitlines()
    rows = [line.split("|") for line in lines]
    max_len = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    return pd.DataFrame(rows)


def load_psv_dfs_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, psv_patterns):
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        dfs = []
        for pat in psv_patterns:
            psv_file = find_psv_by_glob(inner_extract_root, pat)
            dfs.append(read_psv_preserve_shape(psv_file))

        return dfs
    finally:
        tmpdir.cleanup()


# =========================
# Excel helpers
# =========================
def autosize_columns(ws, col_start: int, col_end: int, row_start: int, row_end: int, min_w=10, max_w=60):
    for c in range(col_start, col_end + 1):
        max_len = 0
        for r in range(row_start, row_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, max_len + 2), max_w)


def write_title(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_LEFT


def write_section_label(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = ALIGN_LEFT


def write_df_values_only(ws, df: pd.DataFrame, start_row: int, start_col: int):
    nrows, ncols = df.shape
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + i, column=start_col + j, value=df.iat[i, j])

    top = start_row
    left = start_col
    bottom = start_row + nrows - 1
    right = start_col + ncols - 1
    return top, left, bottom, right


def compare_and_highlight(ws, rng_after, rng_before):
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    a_rows = a_bottom - a_top + 1
    a_cols = a_right - a_left + 1
    b_rows = b_bottom - b_top + 1
    b_cols = b_right - b_left + 1

    max_rows = max(a_rows, b_rows)
    max_cols = max(a_cols, b_cols)

    for r in range(max_rows):
        for c in range(max_cols):
            a_exists = (r < a_rows and c < a_cols)
            b_exists = (r < b_rows and c < b_cols)

            a_r, a_c = a_top + r, a_left + c
            b_r, b_c = b_top + r, b_left + c

            a_val = ws.cell(a_r, a_c).value if a_exists else ""
            b_val = ws.cell(b_r, b_c).value if b_exists else ""

            a_str = "" if a_val is None else str(a_val)
            b_str = "" if b_val is None else str(b_val)

            if a_str != b_str:
                if a_exists:
                    ws.cell(a_r, a_c).fill = DIFF_FILL
                if b_exists:
                    ws.cell(b_r, b_c).fill = DIFF_FILL


def open_or_create_and_clear(filepath: Path, sheet_name: str) -> Workbook:
    """
    If file exists: open workbook, remove/clear all sheets and data.
    Else: create new workbook.
    Ensures one clean sheet with the given name.
    """
    if filepath.exists():
        wb = load_workbook(filepath)
        # Remove all sheets
        for s in list(wb.sheetnames):
            wb.remove(wb[s])
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        # remove default sheet and create named sheet
        default = wb.active
        wb.remove(default)
        ws = wb.create_sheet(sheet_name)

    # Clear sheet fully (extra safety)
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column > 1:
        ws.delete_cols(1, ws.max_column)

    return wb


def build_validation_excel(
    out_file: str,
    sheet_name: str,
    after_tables: list,
    before_tables: list,
    after_section_labels: list,
    before_section_labels: list
):
    out_path = BASE_DIR / out_file
    wb = open_or_create_and_clear(out_path, sheet_name)
    ws = wb[sheet_name]

    # Determine block width (max cols in any table)
    all_tables = after_tables + before_tables
    block_width = max((df.shape[1] for df in all_tables), default=1)

    # Column start positions (match reference)
    left_start_col = AFTER_START_COL
    # Ensure Before starts at least at column J, but also no overlap + 2 blank cols gap
    computed_min_before = left_start_col + block_width + COL_GAP_BETWEEN_BLOCKS
    right_start_col = max(BEFORE_MIN_START_COL, computed_min_before)

    # Titles (row 1)
    write_title(ws, TITLE_ROW, left_start_col, "After_Run Results")
    write_title(ws, TITLE_ROW, right_start_col, "Before_Run Results")

    # Reference label rows
    if len(after_tables) == 3:
        ref_rows = REF_LABEL_ROWS_3
    else:
        ref_rows = REF_LABEL_ROWS_2

    after_ranges = []
    before_ranges = []

    # ---- Write After block ----
    prev_bottom = 0
    for i, df in enumerate(after_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = max(desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES) if prev_bottom else desired_label_row

        write_section_label(ws, label_row, left_start_col, after_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, left_start_col)
        after_ranges.append(rng)
        prev_bottom = rng[2]

    # ---- Write Before block ----
    prev_bottom = 0
    for i, df in enumerate(before_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = max(desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES) if prev_bottom else desired_label_row

        write_section_label(ws, label_row, right_start_col, before_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, right_start_col)
        before_ranges.append(rng)
        prev_bottom = rng[2]

    # ---- Compare and highlight (Context 7) ----
    for i in range(min(len(after_ranges), len(before_ranges))):
        compare_and_highlight(ws, after_ranges[i], before_ranges[i])

    # Autosize columns for both blocks
    last_row = max(after_ranges[-1][2], before_ranges[-1][2]) if after_ranges and before_ranges else ws.max_row
    autosize_columns(ws, left_start_col, left_start_col + block_width - 1, 1, last_row)
    autosize_columns(ws, right_start_col, right_start_col + block_width - 1, 1, last_row)

    wb.save(out_path)


# =========================
# Main Execution
# =========================
def main():
    # ---------------- CNB (2 tables) ----------------
    df_ar_cnb_es, df_ar_cnb_sc = load_psv_dfs_from_run(
        AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )
    df_br_cnb_es, df_br_cnb_sc = load_psv_dfs_from_run(
        BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )

    build_validation_excel(
        out_file="CNB_Validation.xlsx",
        sheet_name="CNB_Validation",
        after_tables=[df_ar_cnb_es, df_ar_cnb_sc],
        before_tables=[df_br_cnb_es, df_br_cnb_sc],
        after_section_labels=[
            "After Run - Error Summary CNB Out (File Name)",
            "After Run - Summary Count CNB Out (File Name)"
        ],
        before_section_labels=[
            "Before Run - Error Summary CNB Out (File Name)",
            "Before Run - Summary Count CNB Out (File Name)"
        ],
    )

    # ---------------- CCMS (3 tables) ----------------
    df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc = load_psv_dfs_from_run(
        AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )
    df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )

    build_validation_excel(
        out_file="CCMS_Validation.xlsx",
        sheet_name="CCMS_Validation",
        after_tables=[df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc],
        before_tables=[df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc],
        after_section_labels=[
            "After Run - Error Summary CCMS Out (File Name)",
            "After Run - Summary CCMS Out (File Name)",
            "After Run - Summary Count CCMS Out (File Name)"
        ],
        before_section_labels=[
            "Before Run - Error Summary CCMS Out (File Name)",
            "Before Run - Summary CCMS Out (File Name)",
            "Before Run - Summary Count CCMS Out (File Name)"
        ],
    )

    # ---------------- CMS (3 tables) ----------------
    df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc = load_psv_dfs_from_run(
        AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )
    df_br_cms_es, df_br_cms_sc, df_br_cms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )

    build_validation_excel(
        out_file="CMS_Validation.xlsx",
        sheet_name="CMS_Validation",
        after_tables=[df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc],
        before_tables=[df_br_cms_es, df_br_cms_sc, df_br_cms_scc],
        after_section_labels=[
            "After Run - Error Summary CMS Out (File Name)",
            "After Run - Summary CMS Out (File Name)",
            "After Run - Summary Count CMS Out (File Name)"
        ],
        before_section_labels=[
            "Before Run - Error Summary CMS Out (File Name)",
            "Before Run - Summary CMS Out (File Name)",
            "Before Run - Summary Count CMS Out (File Name)"
        ],
    )

    print("✅ Completed. Generated/updated:")
    print(" - CNB_Validation.xlsx")
    print(" - CCMS_Validation.xlsx")
    print(" - CMS_Validation.xlsx")
    print("✅ Differences highlighted in light yellow. No other table coloring applied.")


if __name__ == "__main__":
    main()