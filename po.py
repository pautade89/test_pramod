import tarfile
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


"""po.py — extract PSVs from nested TARs and produce Excel validation reports.

This script:
- Locates the latest run TAR files under `After_Run` and `Before_Run`
- Extracts nested inner TARs (e.g., `cnb_out*`, `ccms_out*`)
- Reads pipe-separated PSV files into pandas DataFrames
- Builds Excel workbooks comparing "after" vs "before" tables with
  highlighted differences.

Functions are small and documented to make testing and reuse easier.
"""

# ==========================================================
# Base directory (so script works even if run from elsewhere)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
AFTER_DIR = BASE_DIR / "After_Run"
BEFORE_DIR = BASE_DIR / "Before_Run"


# =========================
# Dynamic PSV filename patterns (Context 2)
# =========================
cnb_list = [
    "error_summary_cnb_out_*.psv",
    "summary_count_cnb_out_*.psv",
]

ccms_list = [
    "error_summary_ccms_out_*.psv",
    "summary_ccms_out_*.psv",
    "summary_count_ccms_out_*.psv",
]

cms_list = [
    "error_summary_cms_out_*.psv",
    "summary_cms_out_*.psv",
    "summary_count_cms_out_*.psv",
]


# =========================
# TAR "prefixes" (timestamp part is dynamic)
# =========================
CNB_OUTER_PREFIX = "cnb_in_out"
CNB_INNER_OUT_PREFIX = "cnb_out"

CCMS_OUTER_PREFIX = "lgd_ccms_in_out"
CCMS_INNER_OUT_PREFIX = "ccms_out"

COMM_OUTER_PREFIX = "lgd_commercial_in_out"
CMS_INNER_OUT_PREFIX = "cms_out"
ESN_INNER_OUT_PREFIX = "esn_out"


# =========================
# Excel Layout (Context 5/6)
# =========================
ROW_GAP_BETWEEN_TABLES = 3
COL_GAP_BETWEEN_BLOCKS = 2

TITLE_ROW = 1
FIRST_SECTION_TITLE_ROW = 3

# Context 7 highlight
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


# =========================
# Robust File Find Helpers
# =========================
TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz")

def list_dir_files(folder: Path):
    """Return a sorted list of files directly under `folder`.

    Returns an empty list if `folder` does not exist.
    """
    if not folder.exists():
        return []
    return sorted([p for p in folder.iterdir() if p.is_file()])


def find_by_prefix(folder: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    """
    Find a file directly under `folder` whose name starts with `prefix` and
    ends with one of `suffixes`. If multiple candidates exist, the most
    recently modified file is returned.

    Raises:
        FileNotFoundError: if `folder` does not exist or no matching files are found.
    """
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")

    candidates = []
    for p in folder.iterdir():
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes):
            candidates.append(p)

    if not candidates:
        # Debug-friendly message
        files_here = [p.name for p in list_dir_files(folder)]
        raise FileNotFoundError(
            f"Couldn't find any file starting with '{prefix}' under: {folder}\n"
            f"Files present: {files_here}"
        )

    # pick most recently modified (helps if multiple runs exist)
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"⚠️ Multiple matches for '{prefix}' in {folder}. Using newest: {candidates[0].name}")
    return candidates[0]


def find_inside_extracted(root: Path, prefix: str, suffixes=TAR_SUFFIXES) -> Path:
    """
    Search recursively under `root` for a file whose name starts with `prefix`
    and ends with one of `suffixes`. Returns the newest match when multiple
    files are found.

    Raises:
        FileNotFoundError: if no matching file is found under `root`.
    """
    matches = []
    for p in root.rglob("*"):
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(suffixes):
            matches.append(p)

    if not matches:
        raise FileNotFoundError(f"Couldn't find inner tar starting with '{prefix}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple inner matches for '{prefix}'. Using newest: {matches[0].name}")
    return matches[0]


def find_psv_by_glob(root: Path, pattern: str) -> Path:
    """
    Find a PSV by wildcard (e.g., error_summary_cnb_out_*.psv) recursively under `root`.
    If multiple matches are found, return the most recently modified one.

    Raises:
        FileNotFoundError: if no matching PSV files are found.
    """
    matches = sorted(root.rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Couldn't find PSV '{pattern}' under extracted: {root}")

    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple PSV matches for '{pattern}'. Using newest: {matches[0].name}")
    return matches[0]


def extract_tar(tar_path: Path, extract_to: Path) -> None:
    """Extract a tar (or compressed tar) archive to `extract_to`.

    Uses tarfile with automatic compression detection ("r:*").
    """
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(path=extract_to)


def read_psv_to_df(psv_path: Path) -> pd.DataFrame:
    """Read a pipe-separated PSV file into a pandas DataFrame.

    All columns are read as strings and missing values are preserved as empty strings.
    """
    return pd.read_csv(
        psv_path,
        sep="|",
        dtype=str,
        keep_default_na=False,
        engine="python"
    )


# =========================
# Generic Loader
# =========================
def load_psv_dfs_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, psv_patterns):
    """
    Locate an outer tar in `run_folder` by `outer_prefix`, extract it, then
    locate and extract one or more inner tars identified by `inner_prefixes`.
    Finally, read PSV files matching `psv_patterns` (wildcards) into DataFrames.

    Returns:
        list[pd.DataFrame]: DataFrames read in the same order as `psv_patterns`.
    """
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        dfs = []
        for pat in psv_patterns:
            psv_file = find_psv_by_glob(inner_extract_root, pat)
            dfs.append(read_psv_to_df(psv_file))

        return dfs
    finally:
        tmpdir.cleanup()


# =========================
# Excel helpers + diff highlight
# =========================
def autosize_columns(ws, col_start: int, col_end: int, row_start: int, row_end: int, min_w=10, max_w=60):
    """Auto-size worksheet columns between col_start..col_end based on the
    text length in rows row_start..row_end. Width is clamped between min_w and max_w.
    """
    for c in range(col_start, col_end + 1):
        max_len = 0
        for r in range(row_start, row_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, max_len + 2), max_w)


def write_title(ws, row: int, col: int, text: str, span_cols: int):
    """Write a bold title to `ws` at (row, col) and optionally merge across columns.

    `span_cols` controls how many columns the title spans (1 = no merge).
    """
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_LEFT
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span_cols - 1)


def write_section_label(ws, row: int, col: int, text: str):
    """Write a bold, left-aligned section label at (row, col)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = ALIGN_LEFT


def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int):
    """Write DataFrame `df` into `ws` starting at (start_row, start_col).

    Returns a tuple (top, left, bottom, right) representing the written range
    (header included).
    """
    nrows, ncols = df.shape

    # Header
    for j, name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(name))
        cell.font = BOLD
        cell.alignment = ALIGN_LEFT

    # Data
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=df.iat[i, j])

    top = start_row
    left = start_col
    bottom = start_row + nrows
    right = start_col + ncols - 1
    return top, left, bottom, right


def compare_and_highlight(ws, rng_after, rng_before):
    """Compare two rectangular ranges written on the same worksheet and highlight differences.

    Each range is a tuple (top, left, bottom, right). Cells present in one range
    but not the other are treated as empty strings for comparison.
    """
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    a_rows = a_bottom - a_top + 1
    a_cols = a_right - a_left + 1
    b_rows = b_bottom - b_top + 1
    b_cols = b_right - b_left + 1

    max_rows = max(a_rows, b_rows)
    max_cols = max(a_cols, b_cols)

    for r in range(max_rows):
        for c in range(max_cols):
            # Determine whether a cell exists in each range
            a_exists = (r < a_rows and c < a_cols)
            b_exists = (r < b_rows and c < b_cols)

            a_r, a_c = a_top + r, a_left + c
            b_r, b_c = b_top + r, b_left + c

            a_val = ws.cell(a_r, a_c).value if a_exists else ""
            b_val = ws.cell(b_r, b_c).value if b_exists else ""

            # Normalize values to strings for comparison
            a_str = "" if a_val is None else str(a_val)
            b_str = "" if b_val is None else str(b_val)

            if a_str != b_str:
                # Highlight any differing cells in either range
                if a_exists:
                    ws.cell(a_r, a_c).fill = DIFF_FILL
                if b_exists:
                    ws.cell(b_r, b_c).fill = DIFF_FILL


def build_validation_excel(out_file, sheet_name, after_tables, before_tables, after_titles, before_titles):
    """Create an Excel file comparing `after_tables` vs `before_tables`.

    - `after_tables` and `before_tables` are lists of DataFrames written in
      parallel blocks (left/right). `after_titles` and `before_titles` are
      corresponding section labels.
    - Differences between paired tables are highlighted using `DIFF_FILL`.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    after_max_cols = max(df.shape[1] for df in after_tables) if after_tables else 1
    before_max_cols = max(df.shape[1] for df in before_tables) if before_tables else 1
    block_width = max(after_max_cols, before_max_cols)

    left_start_col = 1
    right_start_col = left_start_col + block_width + COL_GAP_BETWEEN_BLOCKS  # ensures 2 blank columns gap

    write_title(ws, TITLE_ROW, left_start_col, "After_Run Results", block_width)
    write_title(ws, TITLE_ROW, right_start_col, "Before_Run Results", block_width)

    after_ranges = []
    before_ranges = []

    cur_after_label_row = FIRST_SECTION_TITLE_ROW
    cur_before_label_row = FIRST_SECTION_TITLE_ROW

    for i, df in enumerate(after_tables):
        write_section_label(ws, cur_after_label_row, left_start_col, after_titles[i])
        rng = write_df(ws, df, cur_after_label_row + 1, left_start_col)
        after_ranges.append(rng)
        cur_after_label_row = rng[2] + 1 + ROW_GAP_BETWEEN_TABLES

    for i, df in enumerate(before_tables):
        write_section_label(ws, cur_before_label_row, right_start_col, before_titles[i])
        rng = write_df(ws, df, cur_before_label_row + 1, right_start_col)
        before_ranges.append(rng)
        cur_before_label_row = rng[2] + 1 + ROW_GAP_BETWEEN_TABLES

    # Context 7: highlight diffs
    for i in range(min(len(after_ranges), len(before_ranges))):
        compare_and_highlight(ws, after_ranges[i], before_ranges[i])

    last_row = max(
        after_ranges[-1][2] if after_ranges else 1,
        before_ranges[-1][2] if before_ranges else 1
    )

    autosize_columns(ws, left_start_col, left_start_col + block_width - 1, 1, last_row)
    autosize_columns(ws, right_start_col, right_start_col + block_width - 1, 1, last_row)

    wb.save(out_file)


# =========================
# Main
# =========================
def main():
    """Top-level driver: load PSV DataFrames for each channel and produce Excel reports.

    Workflow:
    - Load CNB, CCMS, and CMS PSV sets from both After_Run and Before_Run
    - Build corresponding validation Excel files with highlighted differences
    """
    # CNB
    df_ar_cnb_es, df_ar_cnb_sc = load_psv_dfs_from_run(
        AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )
    df_br_cnb_es, df_br_cnb_sc = load_psv_dfs_from_run(
        BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
    )

    build_validation_excel(
        "CNB_Validation.excel", "CNB_Validation",
        after_tables=[df_ar_cnb_es, df_ar_cnb_sc],
        before_tables=[df_br_cnb_es, df_br_cnb_sc],
        after_titles=[
            "After Run - Error Summary CNB Out (File Name)",
            "After Run - Summary Count CNB Out (File Name)",
        ],
        before_titles=[
            "Before Run - Error Summary CNB Out (File Name)",
            "Before Run - Summary Count CNB Out (File Name)",
        ]
    )

    # CCMS
    df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc = load_psv_dfs_from_run(
        AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )
    df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
    )

    build_validation_excel(
        "CCMS_Validation.excel", "CCMS_Validation",
        after_tables=[df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc],
        before_tables=[df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc],
        after_titles=[
            "After Run - Error Summary CCMS Out (File Name)",
            "After Run - Summary CCMS Out (File Name)",
            "After Run - Summary Count CCMS Out (File Name)",
        ],
        before_titles=[
            "Before Run - Error Summary CCMS Out (File Name)",
            "Before Run - Summary CCMS Out (File Name)",
            "Before Run - Summary Count CCMS Out (File Name)",
        ]
    )

    # CMS (commercial) — extract both cms_out* and esn_out*
    df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc = load_psv_dfs_from_run(
        AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )
    df_br_cms_es, df_br_cms_sc, df_br_cms_scc = load_psv_dfs_from_run(
        BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
    )

    build_validation_excel(
        "CMS_Validation.excel", "CMS_Validation",
        after_tables=[df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc],
        before_tables=[df_br_cms_es, df_br_cms_sc, df_br_cms_scc],
        after_titles=[
            "After Run - Error Summary CMS Out (File Name)",
            "After Run - Summary CMS Out (File Name)",
            "After Run - Summary Count CMS Out (File Name)",
        ],
        before_titles=[
            "Before Run - Error Summary CMS Out (File Name)",
            "Before Run - Summary CMS Out (File Name)",
            "Before Run - Summary Count CMS Out (File Name)",
        ]
    )

    print("✅ Done. Created:")
    print(" - CNB_Validation.excel")
    print(" - CCMS_Validation.excel")
    print(" - CMS_Validation.excel")
    print("✅ Differences highlighted in light yellow.")


if __name__ == "__main__":
    main()