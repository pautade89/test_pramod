
import tarfile
import tempfile
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==========================================================
# Base directory (works even if you run from elsewhere)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
AFTER_DIR = BASE_DIR / "After_Run"
BEFORE_DIR = BASE_DIR / "Before_Run"


# =========================
# PSV patterns (dynamic date part)
# =========================
cnb_list = [
    "error_summary_cnb_out_*.psv",
    "summary_count_cnb_out_*.psv",
]

ccms_list = [
    "error_summary_ccms_out_*.psv",
    "summary_ccms_out_*.psv",
    "summary_count_ccms_out_*.psv",
]

cms_list = [
    "error_summary_cms_out_*.psv",
    "summary_cms_out_*.psv",
    "summary_count_cms_out_*.psv",
]

# Facility PSV patterns
facility_cnb_pattern = "facility_cnb_out_*.psv"
facility_ccms_pattern = "facility_ccms_out_*.psv"
facility_cms_pattern = "facility_cms_out_*.psv"


# =========================
# TAR prefixes
# =========================
CNB_OUTER_PREFIX = "cnb_in_out"
CNB_INNER_OUT_PREFIX = "cnb_out"

CCMS_OUTER_PREFIX = "lgd_ccms_in_out"
CCMS_INNER_OUT_PREFIX = "ccms_out"

COMM_OUTER_PREFIX = "lgd_commercial_in_out"
CMS_INNER_OUT_PREFIX = "cms_out"
ESN_INNER_OUT_PREFIX = "esn_out"


# =========================
# Excel layout rules
# =========================
AFTER_START_COL = 1          # A
BEFORE_MIN_START_COL = 10    # J
COL_GAP_BETWEEN_BLOCKS = 2   # 2 blank columns between blocks
ROW_GAP_BETWEEN_TABLES = 3   # 3 row gap between tables

TITLE_ROW = 1

# Reference label rows (if tables fit into expected slots)
REF_LABEL_ROWS_2 = [3, 12]
REF_LABEL_ROWS_3 = [3, 12, 20]
REF_LABEL_ROWS_4 = [3, 12, 19, 26]
REF_LABEL_ROWS_5 = [3, 12, 19, 26, 33]


# =========================
# Styles
# =========================
TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

# Run-to-run comparison
MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow

# Business rule mismatch
RULE_DIFF_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red

thin = Side(style="thin", color="000000")
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# =========================
# TAR helpers
# =========================
TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz")


def tar_present_by_prefix(folder: Path, prefix: str) -> bool:
    if not folder.exists():
        return False
    return any(
        p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
        for p in folder.iterdir()
    )


def find_by_prefix(folder: Path, prefix: str) -> Path:
    """Pick newest file under folder that starts with prefix and ends with .tar/.tar.gz/.tgz"""
    candidates = [
        p for p in folder.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
    ]
    if not candidates:
        present = sorted([p.name for p in folder.iterdir() if p.is_file()]) if folder.exists() else []
        raise FileNotFoundError(
            f"Couldn't find any tar starting with '{prefix}' under: {folder}\n"
            f"Files present: {present}"
        )
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"⚠️ Multiple matches for '{prefix}' in {folder}. Using newest: {candidates[0].name}")
    return candidates[0]


def find_inside_extracted(root: Path, prefix: str) -> Path:
    """Find newest inner tar recursively inside extracted outer tar folder."""
    matches = [
        p for p in root.rglob("*")
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
    ]
    if not matches:
        raise FileNotFoundError(f"Couldn't find inner tar starting with '{prefix}' under extracted: {root}")
    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple inner matches for '{prefix}'. Using newest: {matches[0].name}")
    return matches[0]


def find_psv_by_glob(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Couldn't find PSV '{pattern}' under extracted: {root}")
    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple PSV matches for '{pattern}'. Using newest: {matches[0].name}")
    return matches[0]


def extract_tar(tar_path: Path, extract_to: Path) -> None:
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(path=extract_to)


# =========================
# PSV parsing
# =========================

def read_psv_preserve_shape(psv_path: Path) -> pd.DataFrame:
    """Split each line by literal '|', pad rows to max columns. No header inference."""
    lines = psv_path.read_text(encoding="utf-8", errors="replace").splitlines()
    rows = [line.split("|") for line in lines]
    max_len = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    return pd.DataFrame(rows)


def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


def _find_facility_header_line(lines):
    """Detect header line in facility PSV (handles metadata first line)."""
    must_have = {"facilityid"}
    segment_alts = {"finalsegmentid", "segmentid", "segment"}
    rate_alts = {"finallgdrate", "lgdrate", "final_lgd_rate"}

    for idx, line in enumerate(lines[:50]):
        cols = [c.strip() for c in line.split("|") if c.strip() != ""]
        norm_cols = {_norm_col(c) for c in cols}
        if must_have.issubset(norm_cols) and (norm_cols & segment_alts) and (norm_cols & rate_alts):
            return idx

    for idx, line in enumerate(lines[:50]):
        if "FacilityID" in line:
            return idx

    return 0


def read_facility_psv_smart(psv_path: Path) -> pd.DataFrame:
    """Read facility PSV with smart header detection."""
    raw_text = psv_path.read_text(encoding="utf-8", errors="replace")
    lines = raw_text.splitlines()
    header_idx = _find_facility_header_line(lines)

    df = pd.read_csv(
        psv_path,
        sep="|",
        engine="python",
        header=0,
        skiprows=header_idx,
        dtype=str,
        keep_default_na=False,
    )

    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, [c for c in df.columns if not _norm_col(c).startswith("unnamed")]]
    return df


# =========================
# Loaders
# =========================

def load_psv_dfs_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, psv_patterns):
    """Extract outer tar -> extract inner tar(s) -> read PSV(s) (preserve-shape)."""
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        dfs = []
        for pat in psv_patterns:
            psv_file = find_psv_by_glob(inner_extract_root, pat)
            dfs.append(read_psv_preserve_shape(psv_file))

        return dfs
    finally:
        tmpdir.cleanup()



def load_facility_df_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, facility_pattern: str) -> pd.DataFrame:
    """Extract outer tar -> extract inner tar(s) -> read facility PSV."""
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        facility_psv = find_psv_by_glob(inner_extract_root, facility_pattern)
        return read_facility_psv_smart(facility_psv)

    finally:
        tmpdir.cleanup()


# =========================
# Facility pivot builder
# =========================

def _resolve_required_cols(df: pd.DataFrame):
    norm_map = {_norm_col(c): c for c in df.columns}

    facility_keys = ["facilityid", "facility_id", "facility id"]
    segment_keys = [
        "finalsegmentid", "final_segment_id", "final segment id",
        "segmentid", "segment_id", "segment id",
        "segment"
    ]
    rate_keys = [
        "finallgdrate", "final_lgd_rate", "final lgd rate", "lgdrate", "lgd rate"
    ]
    #Final EAD mapping keys
    ead_keys = [
        "finalead", "final_ead", "final ead", "ead"
    ]

    def pick(keys):
        for k in keys:
            nk = _norm_col(k)
            if nk in norm_map:
                return norm_map[nk]
        return None

    facility_col = pick(facility_keys)
    segment_col = pick(segment_keys)
    rate_col = pick(rate_keys)
    ead_col = pick(ead_keys)

    missing = []
    if facility_col is None:
        missing.append("FacilityID (or similar)")
    if segment_col is None:
        missing.append("FinalSegmentID/Segment ID (or similar)")
    if rate_col is None:
        missing.append("FinalLGDRate (or similar)")
    if ead_col is None:
        missing.append("FinalEAD (or similar)")

    if missing:
        raise ValueError(
            "Facility file missing required columns.\n"
            f"Missing: {missing}\n"
            f"Found columns: {list(df.columns)}"
        )

    return facility_col, segment_col, rate_col, ead_col



def build_facility_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """Segment-wise count and average LGD (as whole percent string)."""
    facility_col, segment_col, rate_col, ead_col = _resolve_required_cols(df)

    work = df[[facility_col, segment_col, rate_col, ead_col]].copy()
    work[segment_col] = work[segment_col].astype(str).str.strip()

    rate = work[rate_col].astype(str).str.strip()
    rate = rate.str.replace("%", "", regex=False)
    rate = rate.str.replace(",", "", regex=False)
    rate_num = pd.to_numeric(rate, errors="coerce")

    max_rate = rate_num.max(skipna=True)
    if pd.notna(max_rate) and max_rate <= 1.5:
        rate_num = rate_num * 100.0

    work["__rate_num__"] = rate_num

    # EAD into Numeric
    ead = work[ead_col].astype(str).str.strip()
    ead = ead.str.replace(",", "", regex=False)
    ead_num = pd.to_numeric(ead, errors="coerce").fillna(0)
    work["__ead_num__"] = ead_num

    # Pivot summary by Segment
    pt = (
        work.groupby(segment_col, dropna=False)
        .agg(
            **{
                "Count of FacilityID": (facility_col, "count"),
                "Average of FinalLGDRate": ("__rate_num__", "mean"),
                "Sum of FinalEAD": ("__ead_num__", "sum"),
            }
        )
        .reset_index()
        .rename(columns={segment_col: "Segment ID"})
    )
    # Format avg as whole percent string (same as before)
    pt["Average of FinalLGDRate"] = pt["Average of FinalLGDRate"].map(
        lambda x: "" if pd.isna(x) else f"{round(x):.0f}%"
    )

    # Keep Sum of Final EAD Numeric
    pt["Sum of FinalEAD"] = pt["Sum of FinalEAD"].map(
        lambda x: int(x) if pd.notna(x) and float(x).is_integer() else (0 if pd.isna(x) else float(x))
    )

    # Sort of Segment ID numerically
    def _seg_sort_key(v):
        try:
            return (0, int(float(str(v))))
        except Exception:
            return (1, str(v))

    pt = pt.sort_values(by="Segment ID", key=lambda s: s.map(_seg_sort_key), kind="stable").reset_index(drop=True)

    # Add Total row at end
    total_count = pd.to_numeric(pt["Count of FacilityID"], errors="coerce").fillns(0).sum()
    total_ead = pd.to_numeric(pt["Sum of FinalEAD"], errors="coerce").fillns(0).sum()

    total_row = {
        "Segment ID" : "Total",
        "Count of FacilityID" : int(total_count),
        "Average of FinalLGDRate" : "NA",
        "Sum of FinalEAD" : int(total_ead) if float(total_ead).is_integer() else float(total_ead),
    }

    pt = pd.concat([pt, pd.DataFrame(total_row)], ignore_index=True)

    return pt


def df_with_header_row(df: pd.DataFrame) -> pd.DataFrame:
    """Convert DF to display DF by including header as first row (writer writes values only)."""
    header = [list(df.columns)]
    body = df.astype(object).values.tolist()
    return pd.DataFrame(header + body)


# =========================
# Business Rules (User provided final rules)
# =========================

def business_rule_ccms_cms_df() -> pd.DataFrame:
    """Standard LGD rates segment-wise for CCMS/CMS (Basel IV) as provided."""
    rules = [
        (10, "10%"), (11, "15%"), (12, "40%"), (14, "60%"), (15, "75%"), (16, "35%"),
        (17, "50%"), (18, "45%"), (19, "60%"), (3, "40%"), (1, "55%"), (2, "60%"),
        (13, "30%"), (20, "15%"), (26, "35%"), (25, "55%"), (23, "30%"), (21, "20%"),
        (22, "40%"), (30, "30%"), (65, "40%"), (64, "65%"), (34, "35%"), (33, "40%"),
        (43, "20%"), (44, "25%"), (45, "30%"), (46, "35%"), (50, "40%"), (51, "30%"),
        (52, "20%"), (53, "40%"), (54, "60%"), (60, "6%"), (61, "15%"), (63, "30%"),
        (70, "45%"), (97, "Blended (variable)"), (98, "99%"), (99, "40%"),
    ]
    data = []
    for i, (seg, rate) in enumerate(rules, start=1):
        data.append([i, seg, rate])
    return pd.DataFrame(data, columns=["Sr No", "Segment ID", "LGD Rate"])


def business_rule_ccms_cms_map() -> dict:
    df = business_rule_ccms_cms_df()
    return {int(s): str(r) for s, r in zip(df["Segment ID"], df["LGD Rate"])}


def business_rule_cnb_df() -> pd.DataFrame:
    """CNB standard LGD rates segment-wise (excluding struck values from provided images)."""
    rules = [
        (102, "60%"), (104, "50%"), (106, "45%"), (107, "75%"), (110, "10%"), (111, "15%"),
        (112, "40%"), (120, "15%"), (121, "35%"), (124, "20%"), (127, "55%"), (135, "40%"),
        (143, "20%"), (144, "25%"), (145, "30%"), (146, "35%"), (147, "41%"), (148, "57%"),
        (155, "35%"), (160, "6%"), (164, "15%"), (165, "30%"), (166, "40%"), (170, "45%"),
        (197, "Blended (variable)"), (198, "99%"), (199, "45%"),
    ]
    data = []
    for i, (seg, rate) in enumerate(rules, start=1):
        data.append([i, seg, rate])
    return pd.DataFrame(data, columns=["Sr No", "Segment ID", "LGD Rate"])


def business_rule_cnb_map() -> dict:
    df = business_rule_cnb_df()
    return {int(s): str(r) for s, r in zip(df["Segment ID"], df["LGD Rate"])}


# =========================
# Excel helpers
# =========================

def open_or_create_and_clear(filepath: Path, sheet_name: str):
    if filepath.exists():
        wb = load_workbook(filepath)
        for s in list(wb.sheetnames):
            wb.remove(wb[s])
        wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        wb.create_sheet(sheet_name)
    return wb


def autosize_columns(ws, col_start: int, col_end: int, row_start: int, row_end: int, min_w=10, max_w=60):
    for c in range(col_start, col_end + 1):
        max_len = 0
        for r in range(row_start, row_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, max_len + 2), max_w)


def write_title(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_LEFT


def write_section_label(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = ALIGN_LEFT


def write_df_values_only(ws, df: pd.DataFrame, start_row: int, start_col: int):
    nrows, ncols = df.shape
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + i, column=start_col + j, value=df.iat[i, j])

    top = start_row
    left = start_col
    bottom = start_row + nrows - 1
    right = start_col + ncols - 1
    return top, left, bottom, right


def apply_table_cell_borders(ws, rng):
    """Apply borders AND center alignment only inside the table range."""
    top, left, bottom, right = rng
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            cell = ws.cell(r, c)
            cell.border = CELL_BORDER
            cell.alignment = ALIGN_CENTER


def compare_and_color(ws, rng_after, rng_before):
    """Run-to-run comparison: Match -> green, Diff -> yellow. Apply to both sides."""
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    a_rows = a_bottom - a_top + 1
    a_cols = a_right - a_left + 1
    b_rows = b_bottom - b_top + 1
    b_cols = b_right - b_left + 1

    max_rows = max(a_rows, b_rows)
    max_cols = max(a_cols, b_cols)

    for r in range(max_rows):
        for c in range(max_cols):
            a_exists = (r < a_rows and c < a_cols)
            b_exists = (r < b_rows and c < b_cols)

            a_r, a_c = a_top + r, a_left + c
            b_r, b_c = b_top + r, b_left + c

            a_val = ws.cell(a_r, a_c).value if a_exists else ""
            b_val = ws.cell(b_r, b_c).value if b_exists else ""

            a_str = "" if a_val is None else str(a_val)
            b_str = "" if b_val is None else str(b_val)

            fill = MATCH_FILL if a_str == b_str else DIFF_FILL

            if a_exists:
                ws.cell(a_r, a_c).fill = fill
            if b_exists:
                ws.cell(b_r, b_c).fill = fill


# =========================
# Business Rule Validation Helpers
# =========================

def _parse_segment_id(val):
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _norm_percent_str(val: str) -> str:
    """Normalize percent strings like '60', '60%', '60.0%' -> '60%'."""
    if val is None:
        return ""
    s = str(val).strip()
    if s == "":
        return ""
    if "blended" in s.lower():
        return "Blended (variable)"
    s = s.replace(" ", "")
    if s.endswith("%"):
        s_num = s[:-1]
    else:
        s_num = s
    try:
        n = float(s_num)
        return f"{round(n):.0f}%"
    except Exception:
        return s


def _index_rows_by_segment(ws, rng, header_rows: int, seg_col_offset: int) -> dict:
    """Return {segment_id: absolute_row_index_in_ws} for a table range."""
    top, left, bottom, right = rng
    seg_to_row = {}
    for r in range(top + header_rows, bottom + 1):
        seg = _parse_segment_id(ws.cell(r, left + seg_col_offset).value)
        if seg is None:
            continue
        if seg not in seg_to_row:
            seg_to_row[seg] = r
    return seg_to_row


def apply_pivot_business_rule_validation(
    ws,
    rng_after,
    rng_before,
    rule_map: dict,
    header_rows: int = 1,
    seg_col_offset: int = 0,
    lgd_col_offset: int = 2,
):
    """
    Business Rule validation for pivot tables (CNB / CCMS / CMS).

    FIXED:
    - Validate each side using its OWN Segment ID row (does not assume same ordering).

    Coloring is applied on LGD (Average) cells only:
      - Business rule mismatch => light red
      - If rule matches, preserve run-to-run meaning for SAME Segment ID:
            After vs Before same => light green
            After vs Before different OR missing counterpart => light yellow

    For segments where rule value is 'Blended (variable)', treat as always OK.
    """
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    seg_to_row_after = _index_rows_by_segment(ws, rng_after, header_rows, seg_col_offset)
    seg_to_row_before = _index_rows_by_segment(ws, rng_before, header_rows, seg_col_offset)

    all_segments = set(seg_to_row_after.keys()) | set(seg_to_row_before.keys())

    for seg in all_segments:
        expected = rule_map.get(seg)
        if expected is None:
            continue

        expected_norm = _norm_percent_str(expected)
        blended = (expected_norm == "Blended (variable)")

        # AFTER side
        if seg in seg_to_row_after:
            r_after = seg_to_row_after[seg]
            a_cell = ws.cell(r_after, a_left + lgd_col_offset)
            a_val_norm = _norm_percent_str(a_cell.value)
            a_ok = True if blended else (a_val_norm == expected_norm)

            if seg in seg_to_row_before:
                r_before = seg_to_row_before[seg]
                b_val_norm = _norm_percent_str(ws.cell(r_before, b_left + lgd_col_offset).value)
                ab_same = (a_val_norm == b_val_norm)
            else:
                ab_same = False

            if not a_ok:
                a_cell.fill = RULE_DIFF_FILL
            else:
                a_cell.fill = MATCH_FILL if ab_same else DIFF_FILL

        # BEFORE side
        if seg in seg_to_row_before:
            r_before = seg_to_row_before[seg]
            b_cell = ws.cell(r_before, b_left + lgd_col_offset)
            b_val_norm = _norm_percent_str(b_cell.value)
            b_ok = True if blended else (b_val_norm == expected_norm)

            if seg in seg_to_row_after:
                r_after = seg_to_row_after[seg]
                a_val_norm = _norm_percent_str(ws.cell(r_after, a_left + lgd_col_offset).value)
                ab_same = (a_val_norm == b_val_norm)
            else:
                ab_same = False

            if not b_ok:
                b_cell.fill = RULE_DIFF_FILL
            else:
                b_cell.fill = MATCH_FILL if ab_same else DIFF_FILL


# =========================
# Excel builder
# =========================

def build_validation_excel(
    out_file: str,
    sheet_name: str,
    after_tables: list,
    before_tables: list,
    after_section_labels: list,
    before_section_labels: list,
    rule_checks: list | None = None,
):
    """Build excel with After/Before blocks, run comparison + optional business-rule checks."""
    out_path = BASE_DIR / out_file
    wb = open_or_create_and_clear(out_path, sheet_name)
    ws = wb[sheet_name]

    all_tables = after_tables + before_tables
    block_width = max((df.shape[1] for df in all_tables), default=1)

    left_start_col = AFTER_START_COL
    computed_min_before = left_start_col + block_width + COL_GAP_BETWEEN_BLOCKS
    right_start_col = max(BEFORE_MIN_START_COL, computed_min_before)

    write_title(ws, TITLE_ROW, left_start_col, "After_Run Results")
    write_title(ws, TITLE_ROW, right_start_col, "Before_Run Results")

    if len(after_tables) == 2:
        ref_rows = REF_LABEL_ROWS_2
    elif len(after_tables) == 3:
        ref_rows = REF_LABEL_ROWS_3
    elif len(after_tables) == 4:
        ref_rows = REF_LABEL_ROWS_4
    elif len(after_tables) == 5:
        ref_rows = REF_LABEL_ROWS_5
    else:
        ref_rows = []

    after_ranges = []
    before_ranges = []

    prev_bottom = 0
    for i, df in enumerate(after_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = desired_label_row if prev_bottom == 0 else max(
            desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES
        )

        write_section_label(ws, label_row, left_start_col, after_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, left_start_col)
        after_ranges.append(rng)
        prev_bottom = rng[2]

    prev_bottom = 0
    for i, df in enumerate(before_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = desired_label_row if prev_bottom == 0 else max(
            desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES
        )

        write_section_label(ws, label_row, right_start_col, before_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, right_start_col)
        before_ranges.append(rng)
        prev_bottom = rng[2]

    # Run-to-run compare (green/yellow) - cell-by-cell layout compare
    for i in range(min(len(after_ranges), len(before_ranges))):
        compare_and_color(ws, after_ranges[i], before_ranges[i])

    # Business-rule checks (override LGD cells only: red for rule mismatch)
    if rule_checks:
        for cfg in rule_checks:
            idx = cfg.get("table_index")
            if idx is None:
                continue
            if idx >= len(after_ranges) or idx >= len(before_ranges):
                continue
            apply_pivot_business_rule_validation(
                ws,
                after_ranges[idx],
                before_ranges[idx],
                cfg.get("rule_map", {}),
                header_rows=cfg.get("header_rows", 1),
                seg_col_offset=cfg.get("seg_col_offset", 0),
                lgd_col_offset=cfg.get("lgd_col_offset", 2),
            )

    # Borders + alignment
    for rng in after_ranges:
        apply_table_cell_borders(ws, rng)
    for rng in before_ranges:
        apply_table_cell_borders(ws, rng)

    last_row = max(after_ranges[-1][2], before_ranges[-1][2]) if after_ranges and before_ranges else ws.max_row
    autosize_columns(ws, left_start_col, left_start_col + block_width - 1, 1, last_row)
    autosize_columns(ws, right_start_col, right_start_col + block_width - 1, 1, last_row)

    wb.save(out_path)


# =========================
# Labels
# =========================

def labels_cnb(after=True):
    if after:
        return [
            "After Run - Error Summary CNB Out File",
            "After Run - Summary Count CNB Out File",
            "After Run - Facility CNB Out",
            "Business Rule For CNB",
        ]
    return [
        "Before Run - Error Summary CNB Out File",
        "Before Run - Summary Count CNB Out File",
        "Before Run - Facility CNB Out",
        "Business Rule For CNB",
    ]


def labels_ccms(after=True):
    if after:
        return [
            "After Run - Error Summary CCMS Out File",
            "After Run - Summary CCMS Out File",
            "After Run - Summary Count CCMS Out File",
            "After Run - Facility CCMS Out",
            "Business Rule For CCMS",
        ]
    return [
        "Before Run - Error Summary CCMS Out File",
        "Before Run - Summary CCMS Out File",
        "Before Run - Summary Count CCMS Out File",
        "Before Run - Facility CCMS Out",
        "Business Rule For CCMS",
    ]


def labels_cms(after=True):
    if after:
        return [
            "After Run - Error Summary CMS Out File",
            "After Run - Summary CMS Out File",
            "After Run - Summary Count CMS Out File",
            "After Run - Facility CMS Out",
            "Business Rule For CMS",
        ]
    return [
        "Before Run - Error Summary CMS Out File",
        "Before Run - Summary CMS Out File",
        "Before Run - Summary Count CMS Out File",
        "Before Run - Facility CMS Out",
        "Business Rule For CMS",
    ]


# =========================
# Presence validation
# =========================

def can_generate(prefix: str, excel_name: str) -> bool:
    after_ok = tar_present_by_prefix(AFTER_DIR, prefix)
    before_ok = tar_present_by_prefix(BEFORE_DIR, prefix)

    if after_ok and before_ok:
        return True

    if not after_ok:
        print(f"❌ {prefix}_*.tar tar is absent in After_Run Folder so {excel_name} excel file is not generated.")
    if not before_ok:
        print(f"❌ {prefix}_*.tar tar is absent in Before_Run Folder so {excel_name} excel file is not generated.")
    return False


# =========================
# Main
# =========================

def main():
    # CNB (with Business Rule)
    if can_generate(CNB_OUTER_PREFIX, "CNB_Validation.xlsx"):
        df_ar_cnb_es, df_ar_cnb_sc = load_psv_dfs_from_run(
            AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
        )
        df_br_cnb_es, df_br_cnb_sc = load_psv_dfs_from_run(
            BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
        )

        df_ar_facility_cnb_out = load_facility_df_from_run(
            AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, facility_cnb_pattern
        )
        df_br_facility_cnb_out = load_facility_df_from_run(
            BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, facility_cnb_pattern
        )

        df_ar_facility_cnb_out_pt = build_facility_pivot(df_ar_facility_cnb_out)
        df_br_facility_cnb_out_pt = build_facility_pivot(df_br_facility_cnb_out)

        df_ar_facility_cnb_out_pt_disp = df_with_header_row(df_ar_facility_cnb_out_pt)
        df_br_facility_cnb_out_pt_disp = df_with_header_row(df_br_facility_cnb_out_pt)

        df_rule_cnb_disp = df_with_header_row(business_rule_cnb_df())
        rule_map_cnb = business_rule_cnb_map()

        # Tables: Error Summary, Summary Count, Facility Pivot, Business Rule
        build_validation_excel(
            out_file="CNB_Validation.xlsx",
            sheet_name="CNB_Validation",
            after_tables=[df_ar_cnb_es, df_ar_cnb_sc, df_ar_facility_cnb_out_pt_disp, df_rule_cnb_disp],
            before_tables=[df_br_cnb_es, df_br_cnb_sc, df_br_facility_cnb_out_pt_disp, df_rule_cnb_disp],
            after_section_labels=labels_cnb(after=True),
            before_section_labels=labels_cnb(after=False),
            rule_checks=[{
                'table_index': 2,
                'rule_map': rule_map_cnb,
                'header_rows': 1,
                'seg_col_offset': 0,
                'lgd_col_offset': 2,
            }]
        )
        print("✅ Generated CNB_Validation.xlsx")

    # CCMS (with Business Rule)
    if can_generate(CCMS_OUTER_PREFIX, "CCMS_Validation.xlsx"):
        df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc = load_psv_dfs_from_run(
            AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
        )
        df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc = load_psv_dfs_from_run(
            BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
        )

        df_ar_facility_ccms_out = load_facility_df_from_run(
            AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, facility_ccms_pattern
        )
        df_br_facility_ccms_out = load_facility_df_from_run(
            BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, facility_ccms_pattern
        )

        df_ar_facility_ccms_out_pt = build_facility_pivot(df_ar_facility_ccms_out)
        df_br_facility_ccms_out_pt = build_facility_pivot(df_br_facility_ccms_out)

        df_ar_facility_ccms_out_pt_disp = df_with_header_row(df_ar_facility_ccms_out_pt)
        df_br_facility_ccms_out_pt_disp = df_with_header_row(df_br_facility_ccms_out_pt)

        df_rule_ccms_disp = df_with_header_row(business_rule_ccms_cms_df())
        rule_map_ccms = business_rule_ccms_cms_map()

        build_validation_excel(
            out_file="CCMS_Validation.xlsx",
            sheet_name="CCMS_Validation",
            after_tables=[df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc, df_ar_facility_ccms_out_pt_disp, df_rule_ccms_disp],
            before_tables=[df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc, df_br_facility_ccms_out_pt_disp, df_rule_ccms_disp],
            after_section_labels=labels_ccms(after=True),
            before_section_labels=labels_ccms(after=False),
            rule_checks=[{
                'table_index': 3,
                'rule_map': rule_map_ccms,
                'header_rows': 1,
                'seg_col_offset': 0,
                'lgd_col_offset': 2,
            }]
        )
        print("✅ Generated CCMS_Validation.xlsx")

    # CMS (use same Business Rule as CCMS)
    if can_generate(COMM_OUTER_PREFIX, "CMS_Validation.xlsx"):
        df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc = load_psv_dfs_from_run(
            AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
        )
        df_br_cms_es, df_br_cms_sc, df_br_cms_scc = load_psv_dfs_from_run(
            BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
        )

        df_ar_facility_cms_out = load_facility_df_from_run(
            AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], facility_cms_pattern
        )
        df_br_facility_cms_out = load_facility_df_from_run(
            BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], facility_cms_pattern
        )

        df_ar_facility_cms_out_pt = build_facility_pivot(df_ar_facility_cms_out)
        df_br_facility_cms_out_pt = build_facility_pivot(df_br_facility_cms_out)

        df_ar_facility_cms_out_pt_disp = df_with_header_row(df_ar_facility_cms_out_pt)
        df_br_facility_cms_out_pt_disp = df_with_header_row(df_br_facility_cms_out_pt)

        df_rule_cms_disp = df_with_header_row(business_rule_ccms_cms_df())
        rule_map_cms = business_rule_ccms_cms_map()

        build_validation_excel(
            out_file="CMS_Validation.xlsx",
            sheet_name="CMS_Validation",
            after_tables=[df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc, df_ar_facility_cms_out_pt_disp, df_rule_cms_disp],
            before_tables=[df_br_cms_es, df_br_cms_sc, df_br_cms_scc, df_br_facility_cms_out_pt_disp, df_rule_cms_disp],
            after_section_labels=labels_cms(after=True),
            before_section_labels=labels_cms(after=False),
            rule_checks=[{
                'table_index': 3,
                'rule_map': rule_map_cms,
                'header_rows': 1,
                'seg_col_offset': 0,
                'lgd_col_offset': 2,
            }]
        )
        print("✅ Generated CMS_Validation.xlsx")


if __name__ == "__main__":
    main()