import os
import glob
import tarfile
from io import BytesIO
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# -----------------------------
# Configuration
# -----------------------------
LIGHT_GREEN = PatternFill("solid", fgColor="C6EFCE")  # light green
LIGHT_YELLOW = PatternFill("solid", fgColor="FFF2CC") # light yellow
THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

# Excel layout
TOP_TITLE_ROW = 1     # Excel row 1
SUB_TITLE_ROW = 3     # Excel row 3
TABLE_START_ROW = 5   # Excel row 5 (headers written here)

AFTER_START_COL = 1   # Column A
GAP_COLS = 2          # exactly 2 columns gap
# BEFORE_START_COL computed dynamically based on after-table width


# -----------------------------
# TAR/PSV extraction helpers
# -----------------------------
def find_single_tar(folder, prefix_pattern):
    """
    Return first matching tar path under folder for given prefix pattern like 'cnb_in_out_*.tar'
    If multiple exist, uses lexicographic sort.
    """
    pattern = os.path.join(folder, prefix_pattern)
    matches = sorted(glob.glob(pattern))
    return matches[0] if matches else None


def extract_inner_tar_bytes(outer_tar: tarfile.TarFile, inner_tar_glob: str) -> BytesIO:
    """
    From an open outer tarfile, locate an inner tar member matching inner_tar_glob (glob pattern),
    extract it into BytesIO, and return it.
    """
    import fnmatch
    members = outer_tar.getmembers()
    candidates = [m for m in members if fnmatch.fnmatch(os.path.basename(m.name), inner_tar_glob)]
    if not candidates:
        return None

    # pick first match
    inner_member = candidates[0]
    f = outer_tar.extractfile(inner_member)
    if f is None:
        return None

    data = f.read()
    return BytesIO(data)


def read_psv_from_inner_tar(inner_tar_bytes: BytesIO, psv_glob: str) -> pd.DataFrame:
    """
    Open inner tar from BytesIO and read first PSV that matches psv_glob into a DataFrame.
    """
    import fnmatch

    inner_tar_bytes.seek(0)
    with tarfile.open(fileobj=inner_tar_bytes, mode="r:*") as itar:
        members = itar.getmembers()
        psv_candidates = [m for m in members if fnmatch.fnmatch(os.path.basename(m.name), psv_glob)]
        if not psv_candidates:
            return None

        psv_member = psv_candidates[0]
        f = itar.extractfile(psv_member)
        if f is None:
            return None

        # Pandas can read file-like. PSV uses pipe delimiter.
        df = pd.read_csv(f, sep="|", engine="python")
        return df


def load_facility_df(folder, outer_tar_pattern, inner_tar_pattern, psv_pattern) -> pd.DataFrame:
    """
    Find outer tar, open it, extract inner tar bytes, read PSV from inner tar -> dataframe.
    """
    outer_tar_path = find_single_tar(folder, outer_tar_pattern)
    if not outer_tar_path:
        return None

    with tarfile.open(outer_tar_path, mode="r:*") as otar:
        inner_bytes = extract_inner_tar_bytes(otar, inner_tar_pattern)
        if inner_bytes is None:
            return None
        df = read_psv_from_inner_tar(inner_bytes, psv_pattern)
        return df


def load_cms_facility_df(folder) -> pd.DataFrame:
    """
    For commercial case, facility_cms_out_*.psv can be inside cms_out_*.tar OR esn_out_*.tar.
    So try cms_out first, then esn_out.
    """
    outer_tar_path = find_single_tar(folder, "lgd_commercial_in_out_*.tar")
    if not outer_tar_path:
        return None

    with tarfile.open(outer_tar_path, mode="r:*") as otar:
        # Try cms_out_*.tar
        inner_bytes = extract_inner_tar_bytes(otar, "cms_out_*.tar")
        if inner_bytes is not None:
            df = read_psv_from_inner_tar(inner_bytes, "facility_cms_out_*.psv")
            if df is not None:
                return df

        # Try esn_out_*.tar
        inner_bytes = extract_inner_tar_bytes(otar, "esn_out_*.tar")
        if inner_bytes is not None:
            df = read_psv_from_inner_tar(inner_bytes, "facility_cms_out_*.psv")
            if df is not None:
                return df

    return None


# -----------------------------
# Pivot logic
# -----------------------------
def normalize_lgd_rate(series: pd.Series) -> pd.Series:
    """
    Convert FinalLGDRate to numeric, then normalize to fraction (0-1) if values look like 0-100.
    """
    s = pd.to_numeric(series, errors="coerce")
    non_na = s.dropna()
    if not non_na.empty:
        mx = non_na.max()
        # Heuristic: if values are like 55, 60 etc => divide by 100
        if mx > 1.0 and mx <= 100.0:
            s = s / 100.0
    return s


def make_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pivot output:
    FinalSegmentID as first column
    Count of FacilityID
    Average of FinalLGDRate (numeric fraction; will be formatted as % in Excel)
    """
    required = {"FacilityID", "FinalSegmentID", "FinalLGDRate"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    tmp = df[["FacilityID", "FinalSegmentID", "FinalLGDRate"]].copy()
    tmp["FinalLGDRate"] = normalize_lgd_rate(tmp["FinalLGDRate"])

    pt = (
        tmp.groupby("FinalSegmentID", dropna=False)
           .agg(
               **{
                   "Count of FacilityID": ("FacilityID", "count"),
                   "Average Final LGD Rate": ("FinalLGDRate", "mean"),
               }
           )
           .reset_index()
           .rename(columns={"FinalSegmentID": "Final Segment ID"})
    )

    # sort by Final Segment ID if possible (numeric-like)
    def _safe_sort_key(x):
        try:
            return float(x)
        except Exception:
            return float("inf")

    # If segment IDs are numeric-like strings, sort accordingly; else keep original order
    try:
        pt["_k"] = pt["Final Segment ID"].apply(_safe_sort_key)
        pt = pt.sort_values("_k").drop(columns=["_k"])
    except Exception:
        pass

    return pt


def align_pivots(after_pt: pd.DataFrame, before_pt: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Ensure both pivot tables contain the same set of segment IDs (outer join).
    """
    a = after_pt.set_index("Final Segment ID")
    b = before_pt.set_index("Final Segment ID")

    all_idx = a.index.union(b.index)
    a2 = a.reindex(all_idx)
    b2 = b.reindex(all_idx)

    a2 = a2.reset_index()
    b2 = b2.reset_index()
    return a2, b2


# -----------------------------
# Excel writing + formatting
# -----------------------------
def write_side_by_side_excel(
    out_xlsx: str,
    sheet_name: str,
    after_pt: pd.DataFrame,
    before_pt: pd.DataFrame,
    after_title: str,
    before_title: str
):
    """
    Writes after and before pivot tables side by side with a 2-column gap.
    Adds titles, borders, and comparison highlighting.
    """
    # Align both tables to same segments for clean comparison/highlighting
    after_pt_aligned, before_pt_aligned = align_pivots(after_pt, before_pt)

    # Determine start cols (1-indexed for excel)
    after_start_col = AFTER_START_COL
    after_width = after_pt_aligned.shape[1]  # number of columns
    before_start_col = after_start_col + after_width + GAP_COLS

    # Write using pandas
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # Create sheet
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        # Write tables
        after_pt_aligned.to_excel(
            writer, sheet_name=sheet_name,
            index=False, startrow=TABLE_START_ROW - 1, startcol=after_start_col - 1
        )
        before_pt_aligned.to_excel(
            writer, sheet_name=sheet_name,
            index=False, startrow=TABLE_START_ROW - 1, startcol=before_start_col - 1
        )

    # Apply formatting + highlighting using openpyxl
    wb = load_workbook(out_xlsx)
    ws = wb[sheet_name]

    # Titles
    ws.cell(row=TOP_TITLE_ROW, column=after_start_col, value="After_Run Results").font = TITLE_FONT
    ws.cell(row=TOP_TITLE_ROW, column=before_start_col, value="Before_Run Results").font = TITLE_FONT

    ws.cell(row=SUB_TITLE_ROW, column=after_start_col, value=after_title).font = TITLE_FONT
    ws.cell(row=SUB_TITLE_ROW, column=before_start_col, value=before_title).font = TITLE_FONT

    # Header formatting (row = TABLE_START_ROW)
    header_row = TABLE_START_ROW
    for c in range(after_start_col, after_start_col + after_width):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for c in range(before_start_col, before_start_col + after_width):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Determine table ranges
    n_rows = after_pt_aligned.shape[0]
    table_rows_total = n_rows + 1  # includes header
    after_r1 = header_row
    after_r2 = header_row + table_rows_total - 1
    after_c1 = after_start_col
    after_c2 = after_start_col + after_width - 1

    before_r1 = header_row
    before_r2 = header_row + table_rows_total - 1
    before_c1 = before_start_col
    before_c2 = before_start_col + after_width - 1

    # Percent format for "Average Final LGD Rate" column (3rd column in each table)
    # Columns are: Final Segment ID, Count of FacilityID, Average Final LGD Rate
    avg_col_offset = 2  # 0-based offset: 2 => 3rd column
    after_avg_col = after_c1 + avg_col_offset
    before_avg_col = before_c1 + avg_col_offset

    for r in range(header_row + 1, after_r2 + 1):
        ws.cell(row=r, column=after_avg_col).number_format = "0%"
        ws.cell(row=r, column=before_avg_col).number_format = "0%"

    # Borders around both tables (including header)
    def apply_border(r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = THIN_BORDER

    apply_border(after_r1, after_r2, after_c1, after_c2)
    apply_border(before_r1, before_r2, before_c1, before_c2)

    # Auto column widths (simple heuristic)
    def autosize(c1, c2):
        for c in range(c1, c2 + 1):
            max_len = 0
            col_letter = get_column_letter(c)
            for r in range(1, after_r2 + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)

    autosize(after_c1, after_c2)
    autosize(before_c1, before_c2)

    # Comparison highlighting (data cells only, excluding header)
    # Compare each cell by position since we aligned by Final Segment ID.
    # Count column exact, Avg column tolerance
    tol = 1e-9

    def values_match(v1, v2, is_avg=False):
        if v1 is None and v2 is None:
            return True
        if (v1 is None) != (v2 is None):
            return False

        # handle NaN
        try:
            if pd.isna(v1) and pd.isna(v2):
                return True
            if pd.isna(v1) != pd.isna(v2):
                return False
        except Exception:
            pass

        if is_avg:
            try:
                return abs(float(v1) - float(v2)) <= tol
            except Exception:
                return str(v1) == str(v2)
        else:
            return str(v1) == str(v2)

    # Loop through each data row and each column in the table
    for i in range(n_rows):
        excel_row = header_row + 1 + i

        for j in range(after_width):
            after_cell = ws.cell(row=excel_row, column=after_c1 + j)
            before_cell = ws.cell(row=excel_row, column=before_c1 + j)

            is_avg = (j == avg_col_offset)

            match = values_match(after_cell.value, before_cell.value, is_avg=is_avg)
            fill = LIGHT_GREEN if match else LIGHT_YELLOW

            after_cell.fill = fill
            before_cell.fill = fill

    wb.save(out_xlsx)


# -----------------------------
# Main orchestration
# -----------------------------
def main(after_dir="After_Run", before_dir="Before_Run"):
    # CNB
    cnb_after_tar = find_single_tar(after_dir, "cnb_in_out_*.tar")
    cnb_before_tar = find_single_tar(before_dir, "cnb_in_out_*.tar")

    # CCMS
    ccms_after_tar = find_single_tar(after_dir, "lgd_ccms_in_out_*.tar")
    ccms_before_tar = find_single_tar(before_dir, "lgd_ccms_in_out_*.tar")

    # CMS (Commercial)
    cms_after_tar = find_single_tar(after_dir, "lgd_commercial_in_out_*.tar")
    cms_before_tar = find_single_tar(before_dir, "lgd_commercial_in_out_*.tar")

    # ------------------ CNB ------------------
    if cnb_after_tar and cnb_before_tar:
        df_ar_cnb = load_facility_df(after_dir, "cnb_in_out_*.tar", "cnb_out_*.tar", "facility_cnb_out_*.psv")
        df_br_cnb = load_facility_df(before_dir, "cnb_in_out_*.tar", "cnb_out_*.tar", "facility_cnb_out_*.psv")

        if df_ar_cnb is None or df_br_cnb is None:
            print("CNB: Could not locate inner tar or PSV inside one of the folders. CNB excel not generated.")
        else:
            df_ar_cnb_pt = make_pivot(df_ar_cnb)
            df_br_cnb_pt = make_pivot(df_br_cnb)

            write_side_by_side_excel(
                out_xlsx="CNB_Facility_Validation.xlsx",
                sheet_name="CNB_Facility_Validation",
                after_pt=df_ar_cnb_pt,
                before_pt=df_br_cnb_pt,
                after_title="After Run - Facility CNB Out",
                before_title="Before Run - Facility CNB Out"
            )
            print("Generated: CNB_Facility_Validation.xlsx")
    else:
        if not cnb_after_tar:
            print("cnb_in_out_*.tar is absent in After_Run Folder so CNB_Facility_Validation.xlsx is not generated.")
        if not cnb_before_tar:
            print("cnb_in_out_*.tar is absent in Before_Run Folder so CNB_Facility_Validation.xlsx is not generated.")

    # ------------------ CCMS ------------------
    if ccms_after_tar and ccms_before_tar:
        df_ar_ccms = load_facility_df(after_dir, "lgd_ccms_in_out_*.tar", "ccms_out_*.tar", "facility_ccms_out_*.psv")
        df_br_ccms = load_facility_df(before_dir, "lgd_ccms_in_out_*.tar", "ccms_out_*.tar", "facility_ccms_out_*.psv")

        if df_ar_ccms is None or df_br_ccms is None:
            print("CCMS: Could not locate inner tar or PSV inside one of the folders. CCMS excel not generated.")
        else:
            df_ar_ccms_pt = make_pivot(df_ar_ccms)
            df_br_ccms_pt = make_pivot(df_br_ccms)

            write_side_by_side_excel(
                out_xlsx="CCMS_Facility_Validation.xlsx",
                sheet_name="CCMS_Facility_Validation",
                after_pt=df_ar_ccms_pt,
                before_pt=df_br_ccms_pt,
                after_title="After Run - Facility CCMS Out",
                before_title="Before Run - Facility CCMS Out"
            )
            print("Generated: CCMS_Facility_Validation.xlsx")
    else:
        if not ccms_after_tar:
            print("lgd_ccms_in_out_*.tar is absent in After_Run Folder so CCMS_Facility_Validation.xlsx is not generated.")
        if not ccms_before_tar:
            print("lgd_ccms_in_out_*.tar is absent in Before_Run Folder so CCMS_Facility_Validation.xlsx is not generated.")

    # ------------------ CMS ------------------
    if cms_after_tar and cms_before_tar:
        df_ar_cms = load_cms_facility_df(after_dir)
        df_br_cms = load_cms_facility_df(before_dir)

        if df_ar_cms is None or df_br_cms is None:
            print("CMS: Could not locate cms_out/esn_out inner tar or facility_cms_out PSV. CMS excel not generated.")
        else:
            df_ar_cms_pt = make_pivot(df_ar_cms)
            df_br_cms_pt = make_pivot(df_br_cms)

            write_side_by_side_excel(
                out_xlsx="CMS_Facility_Validation.xlsx",
                sheet_name="CMS_Facility_Validation",
                after_pt=df_ar_cms_pt,
                before_pt=df_br_cms_pt,
                after_title="After Run - Facility CMS Out",
                before_title="Before Run - Facility CMS Out"
            )
            print("Generated: CMS_Facility_Validation.xlsx")
    else:
        if not cms_after_tar:
            print("lgd_commercial_in_out_*.tar is absent in After_Run Folder so CMS_Facility_Validation.xlsx is not generated.")
        if not cms_before_tar:
            print("lgd_commercial_in_out_*.tar is absent in Before_Run Folder so CMS_Facility_Validation.xlsx is not generated.")


if __name__ == "__main__":
    # assumes folders are in current working directory:
    # ./After_Run and ./Before_Run
    main("After_Run", "Before_Run")