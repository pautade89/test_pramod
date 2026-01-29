import tarfile
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==========================================================
# Base directory (works even if you run from elsewhere)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
AFTER_DIR = BASE_DIR / "After_Run"
BEFORE_DIR = BASE_DIR / "Before_Run"


# =========================
# PSV patterns (dynamic date part)
# =========================
cnb_list = [
    "error_summary_cnb_out_*.psv",
    "summary_count_cnb_out_*.psv",
]

ccms_list = [
    "error_summary_ccms_out_*.psv",
    "summary_ccms_out_*.psv",
    "summary_count_ccms_out_*.psv",
]

cms_list = [
    "error_summary_cms_out_*.psv",
    "summary_cms_out_*.psv",
    "summary_count_cms_out_*.psv",
]


# =========================
# TAR prefixes (dynamic timestamp part)
# =========================
CNB_OUTER_PREFIX = "cnb_in_out"
CNB_INNER_OUT_PREFIX = "cnb_out"

CCMS_OUTER_PREFIX = "lgd_ccms_in_out"
CCMS_INNER_OUT_PREFIX = "ccms_out"

COMM_OUTER_PREFIX = "lgd_commercial_in_out"
CMS_INNER_OUT_PREFIX = "cms_out"
ESN_INNER_OUT_PREFIX = "esn_out"


# =========================
# Excel layout rules (reference format)
# =========================
AFTER_START_COL = 1          # A
BEFORE_MIN_START_COL = 10    # J (as per reference)
COL_GAP_BETWEEN_BLOCKS = 2   # 2 blank columns between blocks
ROW_GAP_BETWEEN_TABLES = 3   # safety gap if table is taller than reference slots

TITLE_ROW = 1
REF_LABEL_ROWS_2 = [3, 12]       # CNB
REF_LABEL_ROWS_3 = [3, 12, 19]   # CCMS/CMS


# =========================
# Styles
# =========================
TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Context 8 coloring:
MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow

# Borders only within table ranges (grid like sample)
thin = Side(style="thin", color="000000")
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# =========================
# TAR helpers
# =========================
TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz")


def tar_present_by_prefix(folder: Path, prefix: str) -> bool:
    if not folder.exists():
        return False
    return any(
        p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
        for p in folder.iterdir()
    )


def find_by_prefix(folder: Path, prefix: str) -> Path:
    """
    Pick newest file under folder that starts with prefix and ends with .tar/.tar.gz/.tgz
    """
    candidates = [
        p for p in folder.iterdir()
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
    ]
    if not candidates:
        present = sorted([p.name for p in folder.iterdir() if p.is_file()]) if folder.exists() else []
        raise FileNotFoundError(
            f"Couldn't find any tar starting with '{prefix}' under: {folder}\n"
            f"Files present: {present}"
        )
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"⚠️ Multiple matches for '{prefix}' in {folder}. Using newest: {candidates[0].name}")
    return candidates[0]


def find_inside_extracted(root: Path, prefix: str) -> Path:
    """
    Find newest inner tar recursively inside extracted outer tar folder.
    """
    matches = [
        p for p in root.rglob("*")
        if p.is_file() and p.name.startswith(prefix) and p.name.endswith(TAR_SUFFIXES)
    ]
    if not matches:
        raise FileNotFoundError(f"Couldn't find inner tar starting with '{prefix}' under extracted: {root}")
    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple inner matches for '{prefix}'. Using newest: {matches[0].name}")
    return matches[0]


def find_psv_by_glob(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Couldn't find PSV '{pattern}' under extracted: {root}")
    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    if len(matches) > 1:
        print(f"⚠️ Multiple PSV matches for '{pattern}'. Using newest: {matches[0].name}")
    return matches[0]


def extract_tar(tar_path: Path, extract_to: Path) -> None:
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(path=extract_to)


# =========================
# PSV parsing: preserve shape exactly
# =========================
def read_psv_preserve_shape(psv_path: Path) -> pd.DataFrame:
    """
    Split each line by literal '|', pad rows to max columns.
    No header inference; keeps file rows/cols as-is.
    """
    lines = psv_path.read_text(encoding="utf-8", errors="replace").splitlines()
    rows = [line.split("|") for line in lines]
    max_len = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    return pd.DataFrame(rows)


def load_psv_dfs_from_run(run_folder: Path, outer_prefix: str, inner_prefixes, psv_patterns):
    """
    Extract outer tar by prefix -> extract inner tar(s) by prefix -> read PSV(s) by wildcard.
    """
    outer_tar = find_by_prefix(run_folder, outer_prefix)

    tmpdir = tempfile.TemporaryDirectory()
    tmp_root = Path(tmpdir.name)

    try:
        outer_extract = tmp_root / "outer"
        outer_extract.mkdir(parents=True, exist_ok=True)
        extract_tar(outer_tar, outer_extract)

        inner_extract_root = tmp_root / "inner"
        inner_extract_root.mkdir(parents=True, exist_ok=True)

        if isinstance(inner_prefixes, str):
            inner_prefixes = [inner_prefixes]

        for pref in inner_prefixes:
            inner_tar = find_inside_extracted(outer_extract, pref)
            dest = inner_extract_root / inner_tar.stem
            dest.mkdir(parents=True, exist_ok=True)
            extract_tar(inner_tar, dest)

        dfs = []
        for pat in psv_patterns:
            psv_file = find_psv_by_glob(inner_extract_root, pat)
            dfs.append(read_psv_preserve_shape(psv_file))

        return dfs
    finally:
        tmpdir.cleanup()


# =========================
# Excel helpers
# =========================
def open_or_create_and_clear(filepath: Path, sheet_name: str):
    """
    If file exists -> open workbook -> remove all sheets -> create fresh sheet
    Else -> create new workbook with only that sheet
    """
    if filepath.exists():
        wb = load_workbook(filepath)
        for s in list(wb.sheetnames):
            wb.remove(wb[s])
        wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        wb.create_sheet(sheet_name)
    return wb


def autosize_columns(ws, col_start: int, col_end: int, row_start: int, row_end: int, min_w=10, max_w=60):
    for c in range(col_start, col_end + 1):
        max_len = 0
        for r in range(row_start, row_end + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_w, max_len + 2), max_w)


def write_title(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_LEFT


def write_section_label(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = ALIGN_LEFT


def write_df_values_only(ws, df: pd.DataFrame, start_row: int, start_col: int):
    """
    Write DF values only (no implicit headers).
    Returns range: (top, left, bottom, right)
    """
    nrows, ncols = df.shape
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + i, column=start_col + j, value=df.iat[i, j])

    top = start_row
    left = start_col
    bottom = start_row + nrows - 1
    right = start_col + ncols - 1
    return top, left, bottom, right


def apply_table_cell_borders(ws, rng):
    """
    Apply borders ONLY to cells inside the dataframe/table range (grid like sample).
    No borders outside the table.
    """
    top, left, bottom, right = rng
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            ws.cell(r, c).border = CELL_BORDER


def compare_and_color(ws, rng_after, rng_before):
    """
    Context 8:
    - Match -> light green
    - Diff  -> light yellow
    Apply to BOTH sides.
    """
    a_top, a_left, a_bottom, a_right = rng_after
    b_top, b_left, b_bottom, b_right = rng_before

    a_rows = a_bottom - a_top + 1
    a_cols = a_right - a_left + 1
    b_rows = b_bottom - b_top + 1
    b_cols = b_right - b_left + 1

    max_rows = max(a_rows, b_rows)
    max_cols = max(a_cols, b_cols)

    for r in range(max_rows):
        for c in range(max_cols):
            a_exists = (r < a_rows and c < a_cols)
            b_exists = (r < b_rows and c < b_cols)

            a_r, a_c = a_top + r, a_left + c
            b_r, b_c = b_top + r, b_left + c

            a_val = ws.cell(a_r, a_c).value if a_exists else ""
            b_val = ws.cell(b_r, b_c).value if b_exists else ""

            a_str = "" if a_val is None else str(a_val)
            b_str = "" if b_val is None else str(b_val)

            fill = MATCH_FILL if a_str == b_str else DIFF_FILL

            if a_exists:
                ws.cell(a_r, a_c).fill = fill
            if b_exists:
                ws.cell(b_r, b_c).fill = fill


def build_validation_excel(
    out_file: str,
    sheet_name: str,
    after_tables: list,
    before_tables: list,
    after_section_labels: list,
    before_section_labels: list
):
    out_path = BASE_DIR / out_file
    wb = open_or_create_and_clear(out_path, sheet_name)
    ws = wb[sheet_name]

    # Determine maximum block width to calculate Before start col safely
    all_tables = after_tables + before_tables
    block_width = max((df.shape[1] for df in all_tables), default=1)

    left_start_col = AFTER_START_COL
    computed_min_before = left_start_col + block_width + COL_GAP_BETWEEN_BLOCKS
    right_start_col = max(BEFORE_MIN_START_COL, computed_min_before)

    # Titles
    write_title(ws, TITLE_ROW, left_start_col, "After_Run Results")
    write_title(ws, TITLE_ROW, right_start_col, "Before_Run Results")

    # Reference label rows by number of tables
    ref_rows = REF_LABEL_ROWS_3 if len(after_tables) == 3 else REF_LABEL_ROWS_2

    after_ranges = []
    before_ranges = []

    # Write After block with reference spacing
    prev_bottom = 0
    for i, df in enumerate(after_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = desired_label_row if prev_bottom == 0 else max(desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)

        write_section_label(ws, label_row, left_start_col, after_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, left_start_col)
        after_ranges.append(rng)
        prev_bottom = rng[2]

    # Write Before block with reference spacing
    prev_bottom = 0
    for i, df in enumerate(before_tables):
        desired_label_row = ref_rows[i] if i < len(ref_rows) else (prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)
        label_row = desired_label_row if prev_bottom == 0 else max(desired_label_row, prev_bottom + 1 + ROW_GAP_BETWEEN_TABLES)

        write_section_label(ws, label_row, right_start_col, before_section_labels[i])
        rng = write_df_values_only(ws, df, label_row + 1, right_start_col)
        before_ranges.append(rng)
        prev_bottom = rng[2]

    # Compare & color + apply borders only within each table range
    for i in range(min(len(after_ranges), len(before_ranges))):
        compare_and_color(ws, after_ranges[i], before_ranges[i])

    # Apply borders to each dataframe/table range (ONLY tables)
    for rng in after_ranges:
        apply_table_cell_borders(ws, rng)
    for rng in before_ranges:
        apply_table_cell_borders(ws, rng)

    # Autosize columns for both blocks
    last_row = max(after_ranges[-1][2], before_ranges[-1][2]) if after_ranges and before_ranges else ws.max_row
    autosize_columns(ws, left_start_col, left_start_col + block_width - 1, 1, last_row)
    autosize_columns(ws, right_start_col, right_start_col + block_width - 1, 1, last_row)

    wb.save(out_path)


# =========================
# Final labels: '(File Name)' -> 'File'
# =========================
def labels_cnb(after=True):
    if after:
        return [
            "After Run - Error Summary CNB Out File",
            "After Run - Summary Count CNB Out File",
        ]
    return [
        "Before Run - Error Summary CNB Out File",
        "Before Run - Summary Count CNB Out File",
    ]


def labels_ccms(after=True):
    if after:
        return [
            "After Run - Error Summary CCMS Out File",
            "After Run - Summary CCMS Out File",
            "After Run - Summary Count CCMS Out File",
        ]
    return [
        "Before Run - Error Summary CCMS Out File",
        "Before Run - Summary CCMS Out File",
        "Before Run - Summary Count CCMS Out File",
    ]


def labels_cms(after=True):
    if after:
        return [
            "After Run - Error Summary CMS Out File",
            "After Run - Summary CMS Out File",
            "After Run - Summary Count CMS Out File",
        ]
    return [
        "Before Run - Error Summary CMS Out File",
        "Before Run - Summary CMS Out File",
        "Before Run - Summary Count CMS Out File",
    ]


# =========================
# NEW: Context 8 (After + Before presence validation)
# =========================
def can_generate(prefix: str, excel_name: str) -> bool:
    """
    Generate report ONLY if the outer tar exists in BOTH After_Run and Before_Run.
    If missing, print required terminal message(s).
    """
    after_ok = tar_present_by_prefix(AFTER_DIR, prefix)
    before_ok = tar_present_by_prefix(BEFORE_DIR, prefix)

    if after_ok and before_ok:
        return True

    # Print missing messages as per your convention
    if not after_ok:
        print(f"❌ {prefix}_*.tar tar is absent in After_Run Folder so {excel_name} excel file is not generated.")
    if not before_ok:
        print(f"❌ {prefix}_*.tar tar is absent in Before_Run Folder so {excel_name} excel file is not generated.")
    return False


# =========================
# Main
# =========================
def main():
    # CNB: generate only if present in BOTH folders
    if can_generate(CNB_OUTER_PREFIX, "CNB_Validation.xlsx"):
        df_ar_cnb_es, df_ar_cnb_sc = load_psv_dfs_from_run(
            AFTER_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
        )
        df_br_cnb_es, df_br_cnb_sc = load_psv_dfs_from_run(
            BEFORE_DIR, CNB_OUTER_PREFIX, CNB_INNER_OUT_PREFIX, cnb_list
        )

        build_validation_excel(
            out_file="CNB_Validation.xlsx",
            sheet_name="CNB_Validation",
            after_tables=[df_ar_cnb_es, df_ar_cnb_sc],
            before_tables=[df_br_cnb_es, df_br_cnb_sc],
            after_section_labels=labels_cnb(after=True),
            before_section_labels=labels_cnb(after=False),
        )
        print("✅ Generated CNB_Validation.xlsx")

    # CCMS: generate only if present in BOTH folders
    if can_generate(CCMS_OUTER_PREFIX, "CCMS_Validation.xlsx"):
        df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc = load_psv_dfs_from_run(
            AFTER_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
        )
        df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc = load_psv_dfs_from_run(
            BEFORE_DIR, CCMS_OUTER_PREFIX, CCMS_INNER_OUT_PREFIX, ccms_list
        )

        build_validation_excel(
            out_file="CCMS_Validation.xlsx",
            sheet_name="CCMS_Validation",
            after_tables=[df_ar_ccms_es, df_ar_ccms_sc, df_ar_ccms_scc],
            before_tables=[df_br_ccms_es, df_br_ccms_sc, df_br_ccms_scc],
            after_section_labels=labels_ccms(after=True),
            before_section_labels=labels_ccms(after=False),
        )
        print("✅ Generated CCMS_Validation.xlsx")

    # CMS (Commercial): generate only if present in BOTH folders
    if can_generate(COMM_OUTER_PREFIX, "CMS_Validation.xlsx"):
        df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc = load_psv_dfs_from_run(
            AFTER_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
        )
        df_br_cms_es, df_br_cms_sc, df_br_cms_scc = load_psv_dfs_from_run(
            BEFORE_DIR, COMM_OUTER_PREFIX, [CMS_INNER_OUT_PREFIX, ESN_INNER_OUT_PREFIX], cms_list
        )

        build_validation_excel(
            out_file="CMS_Validation.xlsx",
            sheet_name="CMS_Validation",
            after_tables=[df_ar_cms_es, df_ar_cms_sc, df_ar_cms_scc],
            before_tables=[df_br_cms_es, df_br_cms_sc, df_br_cms_scc],
            after_section_labels=labels_cms(after=True),
            before_section_labels=labels_cms(after=False),
        )
        print("✅ Generated CMS_Validation.xlsx")


if __name__ == "__main__":
    main()